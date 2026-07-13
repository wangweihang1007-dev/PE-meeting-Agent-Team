from __future__ import annotations

import argparse
import copy
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


FIELDS = [
    "序号",
    "项目简称",
    "成立时间",
    "城市",
    "上市申报预期",
    "前一轮融资时点和估值",
    "已投机构",
    "本轮投前估值",
    "本次融资额（或投后估值）",
    "本次融资截止时间",
    "主营业务",
    "价值",
    "收入",
    "利润",
    "项目来源和录入时间",
    "备注",
    "否决原因",
]


def parse_date(value):
    if not isinstance(value, str):
        return value
    text = value.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return value


def is_formal_row(ws, row: int) -> bool:
    seq = ws.cell(row, 1).value
    name = ws.cell(row, 2).value
    if not name:
        return False
    if isinstance(seq, int):
        return True
    return str(seq).strip().isdigit()


def find_last_formal_row(ws) -> int:
    last = 0
    for row in range(1, ws.max_row + 1):
        if is_formal_row(ws, row):
            last = row
    if last == 0:
        raise ValueError("未找到可复制格式的正式项目行")
    return last


def next_sequence(ws, last_row: int) -> int:
    seqs = []
    for row in range(1, last_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, int):
            seqs.append(value)
        elif isinstance(value, str) and value.strip().isdigit():
            seqs.append(int(value.strip()))
    return (max(seqs) if seqs else 0) + 1


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy.copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy.copy(src.alignment)
        if src.protection:
            dst.protection = copy.copy(src.protection)


def main() -> None:
    parser = argparse.ArgumentParser(description="把项目录入 JSON 追加写入新版 A:Q 项目表")
    parser.add_argument("draft", type=Path)
    parser.add_argument("--workbook", type=Path, default=Path(__file__).resolve().parents[1] / "assets" / "项目表-录入参考指引.xlsx")
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--sheet", default=None)
    args = parser.parse_args()

    data = json.loads(args.draft.read_text(encoding="utf-8-sig"))
    workbook = load_workbook(args.workbook)
    ws = workbook[args.sheet] if args.sheet else workbook.worksheets[0]

    last_row = find_last_formal_row(ws)
    target_row = last_row + 1
    copy_row_style(ws, last_row, target_row, len(FIELDS))

    values = []
    for field in FIELDS:
        if field == "序号":
            values.append(next_sequence(ws, last_row))
        elif field == "成立时间":
            values.append(parse_date(data.get(field, "")))
        else:
            values.append(data.get(field, ""))

    for col, value in enumerate(values, 1):
        ws.cell(target_row, col).value = value

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(json.dumps({"output": str(args.output.resolve()), "row": target_row, "columns": "A:Q"}, ensure_ascii=False))


if __name__ == "__main__":
    main()
